"""
Main FastAPI application
Production Monitoring System for Wire Manufacturing
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
from collections import defaultdict
import json
from fastapi import FastAPI, Depends, HTTPException, Query, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import date, timedelta, datetime
from pydantic import BaseModel, Field

from database import get_db, init_db, SessionLocal
from models import ProductionRecord, StageConfiguration, StageEnum, ShiftEnum
from inventory_models import StageInventory, InventoryTransaction, MaterialMovement
from order_models import ProductionOrder, OrderStageProgress, BatchTracking, OrderStatus, BatchJourneyEvent
from schemas import (
    ProductionRecordCreate, ProductionRecordResponse,
    DashboardSummary, StageDetailResponse, StageStats,
    ProcessFlowNode, AlertsResponse, Alert
)
from analytics import ProductionAnalytics
from inventory_manager import InventoryManager
from inventory_manager import InventoryManager
from config import settings

from user_models import User, UserRole
from auth import get_password_hash
import user_routes
import final_stage_routes

# Initialize FastAPI app
app = FastAPI(
    title=settings.app_name,
    version=settings.app_version,
    description="Real-time production monitoring for wire manufacturing factory"
)

# Include Routers
app.include_router(user_routes.router, prefix="/api", tags=["Authentication"])
app.include_router(final_stage_routes.router, prefix="/api", tags=["Final Stages"])

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.cors_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Initialize database on startup
@app.on_event("startup")
async def startup_event():
    """Initialize database tables and inventory on startup"""
    init_db()
    
    # Initialize inventory tracking and default users
    db = SessionLocal()
    try:
        inventory_mgr = InventoryManager(db)
        inventory_mgr.initialize_inventory()
        
        # Create default admin user if not exists
        admin_user = db.query(User).filter(User.username == "admin").first()
        if not admin_user:
            print("Creating default admin user...")
            hashed_password = get_password_hash("admin123")
            new_admin = User(
                username="admin",
                full_name="System Administrator",
                hashed_password=hashed_password,
                role=UserRole.ADMIN
            )
            db.add(new_admin)
            db.commit()
            print("✅ Default admin user created (admin/admin123)")
            
        # Create default operator user if not exists
        operator_user = db.query(User).filter(User.username == "operator").first()
        if not operator_user:
            print("Creating default operator user...")
            hashed_password = get_password_hash("operator123")
            new_operator = User(
                username="operator",
                full_name="Machine Operator",
                hashed_password=hashed_password,
                role=UserRole.OPERATOR
            )
            db.add(new_operator)
            db.commit()
            print("✅ Default operator user created (operator/operator123)")
            
    finally:
        db.close()

def _update_order_stage_progress(
    db: Session,
    order_id: int,
    stage: StageEnum,
    quantity: float
) -> dict:
    """Update order and stage progress metrics when material is recorded"""
    order = db.query(ProductionOrder).filter(ProductionOrder.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail=f"Order {order_id} not found")

    stage_progress = db.query(OrderStageProgress).filter(
        OrderStageProgress.order_id == order_id,
        OrderStageProgress.stage == stage.value
    ).first()

    if not stage_progress:
        stage_progress = OrderStageProgress(
            order_id=order_id,
            stage=stage.value
        )
        db.add(stage_progress)

    if stage_progress.stage_status == "PENDING":
        stage_progress.stage_status = "IN_PROGRESS"
        stage_progress.started_at = datetime.utcnow()

    stage_progress.output_quantity = (stage_progress.output_quantity or 0.0) + quantity

    if stage_progress.output_quantity >= order.ordered_quantity:
        stage_progress.stage_status = "COMPLETED"
        stage_progress.completed_at = datetime.utcnow()

    order.current_stage = stage.value

    if stage == StageEnum.REWIND:
        order.completed_quantity = (order.completed_quantity or 0.0) + quantity
        if order.completed_quantity >= order.ordered_quantity:
            order.status = OrderStatus.COMPLETED.value
            order.actual_delivery_date = date.today()
        else:
            order.status = OrderStatus.IN_PROGRESS.value
    else:
        if order.status == OrderStatus.PENDING.value:
            order.status = OrderStatus.IN_PROGRESS.value

    db.commit()
    db.refresh(order)

    return {
        "order_id": order.id,
        "order_number": order.order_number,
        "status": order.status,
        "current_stage": order.current_stage,
        "completed_quantity": order.completed_quantity,
        "ordered_quantity": order.ordered_quantity,
        "stage_status": stage_progress.stage_status,
        "stage_output": stage_progress.output_quantity
    }


class OrderCreatePayload(BaseModel):
    """Payload to create a new production order"""
    order_number: str
    customer_name: str
    product_specification: Optional[str] = None
    ordered_quantity: float = Field(gt=0, description="Total quantity ordered in kg")
    target_wire_size_mm: Optional[float] = None
    expected_delivery_date: Optional[date] = None
    priority: int = Field(default=1, ge=1, le=3, description="1=Normal, 2=High, 3=Urgent")
    notes: Optional[str] = None

class BatchCreatePayload(BaseModel):
    """Payload to register a new raw-material coil/batch"""

    batch_number: str
    quantity: float = Field(gt=0, description="Initial batch quantity in kg")
    lot_number: Optional[str] = None
    material_type: Optional[str] = None
    initial_stage: Optional[StageEnum] = StageEnum.RBD
    supplier_name: Optional[str] = None
    received_date: Optional[date] = None
    order_id: Optional[int] = None
    notes: Optional[str] = None


class BatchMovePayload(BaseModel):
    """Payload to move a batch to the next production stage"""

    to_stage: StageEnum
    quantity: float = Field(gt=0, description="Quantity moving forward in kg")
    scrap_quantity: float = Field(default=0.0, ge=0)
    operator: Optional[str] = None
    notes: Optional[str] = None


class BatchHoldPayload(BaseModel):
    """Toggle hold/resume state for a batch"""

    hold: bool = Field(description="True to pause the batch, False to resume")
    reason: Optional[str] = Field(default=None, description="Optional note on why the status changed")


def _stage_sequence() -> List[str]:
    """Ordered list of production stages"""
    return [stage.value for stage in StageEnum]


def _get_next_stage(current_stage: Optional[str], stage_sequence: List[str]) -> Optional[str]:
    if not stage_sequence:
        return None
    if not current_stage:
        return stage_sequence[0]
    if current_stage not in stage_sequence:
        return None
    idx = stage_sequence.index(current_stage)
    return stage_sequence[idx + 1] if idx + 1 < len(stage_sequence) else None


def _serialize_journey_event(event: BatchJourneyEvent) -> dict:
    return {
        "id": event.id,
        "from_stage": event.from_stage,
        "to_stage": event.to_stage,
        "quantity": event.quantity,
        "scrap_quantity": event.scrap_quantity,
        "operator": event.operator,
        "notes": event.notes,
        "movement_date": event.movement_date.isoformat() if event.movement_date else None,
        "created_at": event.created_at.isoformat() if event.created_at else None
    }


def _serialize_batch(
    batch: BatchTracking,
    stage_sequence: List[str],
    order_lookup: Optional[dict] = None,
    last_event: Optional[BatchJourneyEvent] = None
) -> dict:
    """Return a lightweight dict for UI consumption"""
    order_lookup = order_lookup or {}
    current_stage = batch.current_stage
    completed_stages = 0
    if current_stage and current_stage in stage_sequence:
        completed_stages = stage_sequence.index(current_stage) + 1
    progress_pct = round((completed_stages / len(stage_sequence) * 100), 1) if stage_sequence else 0.0
    order = order_lookup.get(batch.order_id)
    hold_history = _parse_hold_history(batch.hold_history or batch.quality_notes)
    latest_hold = hold_history[-1] if hold_history else None

    return {
        "id": batch.id,
        "batch_number": batch.batch_number,
        "lot_number": batch.lot_number,
        "material_type": batch.material_type,
        "quantity": batch.quantity,
        "remaining_quantity": batch.remaining_quantity,
        "current_stage": batch.current_stage,
        "current_status": batch.current_status,
        "supplier_name": batch.supplier_name,
        "received_date": batch.received_date.isoformat() if batch.received_date else None,
        "quality_status": batch.quality_status,
        "order_id": batch.order_id,
        "order_number": order.order_number if order else None,
        "customer_name": order.customer_name if order else None,
        "journey_progress": {
            "completed": completed_stages,
            "total": len(stage_sequence),
            "percentage": progress_pct
        },
        "last_movement": _serialize_journey_event(last_event) if last_event else None,
        "stage_sequence": stage_sequence,
        "hold_history": hold_history,
        "latest_hold": latest_hold
    }


def _parse_hold_history(raw_notes: Optional[str]) -> List[dict]:
    if not raw_notes:
        return []
    try:
        parsed = json.loads(raw_notes)
        if isinstance(parsed, list):
            return parsed
    except (json.JSONDecodeError, TypeError):
        return []
    return []
@app.get("/")
def root():
    """API health check"""
    return {
        "message": "Wire Manufacturing Production Monitor API",
        "version": settings.app_version,
        "status": "operational"
    }

@app.get("/api/summary", response_model=DashboardSummary)
def get_dashboard_summary(
    start_date: Optional[date] = Query(None, description="Start date filter"),
    end_date: Optional[date] = Query(None, description="End date filter"),
    db: Session = Depends(get_db)
):
    """
    Get overall dashboard summary with key metrics
    - Total production (final output from Rewind stage)
    - Total scrap across all stages
    - Overall efficiency
    - Bottleneck stage identification
    - Active alerts count
    """
    analytics = ProductionAnalytics(db)
    
    # Default to last 30 days if not specified
    if not end_date:
        end_date = date.today()
    if not start_date:
        start_date = end_date - timedelta(days=30)
    
    # Get overall metrics
    metrics = analytics.get_overall_metrics(start_date, end_date)
    
    # Detect bottleneck
    bottleneck = analytics.detect_bottleneck(start_date, end_date)
    
    # Get alerts
    alerts = analytics.generate_alerts(start_date, end_date)
    active_alerts = len([a for a in alerts if a.severity in ["warning", "critical"]])
    
    return DashboardSummary(
        total_production=metrics["total_production"],
        total_scrap=metrics["total_scrap"],
        overall_efficiency=metrics["overall_efficiency"],
        bottleneck_stage=bottleneck,
        active_alerts=active_alerts,
        date_range=f"{start_date} to {end_date}"
    )

@app.get("/api/analytics/efficiency")
def get_efficiency_stats(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: Session = Depends(get_db)
):
    """Get daily efficiency statistics"""
    analytics = ProductionAnalytics(db)
    return analytics.get_efficiency_stats(start_date, end_date)

@app.get("/api/analytics/scrap")
def get_scrap_analysis(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: Session = Depends(get_db)
):
    """Get scrap analysis by stage"""
    analytics = ProductionAnalytics(db)
    return analytics.get_scrap_analysis(start_date, end_date)

@app.get("/api/process-flow", response_model=List[ProcessFlowNode])
def get_process_flow(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: Session = Depends(get_db)
):
    """
    Get complete production process flow showing all stages
    Sequential flow: RBD → Inter → Oven → DPC → Rewind
    
    Each node shows:
    - Input/Output quantities
    - Efficiency percentage
    - Status (good/warning/critical)
    - Expected wire sizes
    """
    analytics = ProductionAnalytics(db)
    
    if not end_date:
        end_date = date.today()
    if not start_date:
        start_date = end_date - timedelta(days=30)
    
    return analytics.get_process_flow(start_date, end_date)

@app.get("/api/stage/{stage_name}", response_model=StageDetailResponse)
def get_stage_details(
    stage_name: StageEnum,
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    shift: Optional[ShiftEnum] = Query(None),
    db: Session = Depends(get_db)
):
    """
    Get detailed information for a specific production stage
    
    Returns:
    - Stage statistics (total input/output, efficiency, loss)
    - Recent production records
    - Daily trend data
    """
    analytics = ProductionAnalytics(db)
    
    if not end_date:
        end_date = date.today()
    if not start_date:
        start_date = end_date - timedelta(days=30)
    
    # Get stage statistics
    stats = analytics.get_stage_statistics(stage_name, start_date, end_date, shift)
    
    # Get recent records
    query = db.query(ProductionRecord).filter(
        ProductionRecord.stage == stage_name,
        ProductionRecord.date >= start_date,
        ProductionRecord.date <= end_date
    )
    
    if shift:
        query = query.filter(ProductionRecord.shift == shift)
    
    recent_records = query.order_by(ProductionRecord.date.desc()).limit(50).all()
    
    # Get daily trend
    daily_trend = analytics.get_timeline_data(start_date, end_date, stage_name)
    
    return StageDetailResponse(
        stage=stage_name,
        stats=stats,
        recent_records=recent_records,
        daily_trend=daily_trend
    )

@app.get("/api/timeline")
def get_timeline(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    stage: Optional[StageEnum] = Query(None),
    db: Session = Depends(get_db)
):
    """
    Get production timeline data for charts
    Daily output and efficiency trends
    """
    analytics = ProductionAnalytics(db)
    
    if not end_date:
        end_date = date.today()
    if not start_date:
        start_date = end_date - timedelta(days=30)
    
    timeline = analytics.get_timeline_data(start_date, end_date, stage)
    
    return {
        "timeline": timeline,
        "date_range": {
            "start": start_date.isoformat(),
            "end": end_date.isoformat()
        }
    }

@app.get("/api/alerts", response_model=AlertsResponse)
def get_alerts(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    severity: Optional[str] = Query(None, description="Filter by severity: critical, warning, info"),
    db: Session = Depends(get_db)
):
    """
    Get production alerts
    
    Alert types:
    - Low efficiency warnings
    - High loss/scrap alerts
    - Bottleneck identification
    """
    analytics = ProductionAnalytics(db)
    
    if not end_date:
        end_date = date.today()
    if not start_date:
        start_date = end_date - timedelta(days=7)
    
    alerts = analytics.generate_alerts(start_date, end_date)
    
    # Filter by severity if specified
    if severity:
        alerts = [a for a in alerts if a.severity == severity]
    
    return AlertsResponse(
        alerts=alerts,
        total_count=len(alerts)
    )

@app.post("/api/records", response_model=ProductionRecordResponse)
def create_production_record(
    record: ProductionRecordCreate,
    db: Session = Depends(get_db)
):
    """
    Create a new production record
    Automatically calculates efficiency and loss percentage
    """
    # Calculate efficiency and loss
    efficiency = (record.output_qty / record.input_qty * 100) if record.input_qty > 0 else 0
    loss_percentage = (record.scrap_qty / record.input_qty * 100) if record.input_qty > 0 else 0
    
    db_record = ProductionRecord(
        **record.model_dump(),
        efficiency=efficiency,
        loss_percentage=loss_percentage
    )
    
    db.add(db_record)
    db.commit()
    db.refresh(db_record)
    
    return db_record

@app.get("/api/records", response_model=List[ProductionRecordResponse])
def get_production_records(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    stage: Optional[StageEnum] = Query(None),
    shift: Optional[ShiftEnum] = Query(None),
    db: Session = Depends(get_db)
):
    """
    Get production records with filters
    Supports pagination and filtering by date, stage, and shift
    """
    query = db.query(ProductionRecord)
    
    if start_date:
        query = query.filter(ProductionRecord.date >= start_date)
    if end_date:
        query = query.filter(ProductionRecord.date <= end_date)
    if stage:
        query = query.filter(ProductionRecord.stage == stage)
    if shift:
        query = query.filter(ProductionRecord.shift == shift)
    
    records = query.order_by(ProductionRecord.date.desc()).offset(skip).limit(limit).all()
    
    return records

@app.get("/api/wip")
def get_wip_analysis(
    target_date: date = Query(default=date.today()),
    db: Session = Depends(get_db)
):
    """
    Get Work-in-Progress analysis between stages
    Shows material queued between sequential stages
    """
    analytics = ProductionAnalytics(db)
    
    wip_data = []
    
    # Calculate WIP between each sequential stage
    stage_pairs = [
        (StageEnum.RBD, StageEnum.INTER, "RBD → Inter"),
        (StageEnum.INTER, StageEnum.OVEN, "Inter → Oven"),
        (StageEnum.OVEN, StageEnum.DPC, "Oven → DPC"),
        (StageEnum.DPC, StageEnum.REWIND, "DPC → Rewind")
    ]
    
    for from_stage, to_stage, label in stage_pairs:
        wip = analytics.calculate_wip(from_stage, to_stage, target_date)
        wip_data.append({
            "from_stage": from_stage.value,
            "to_stage": to_stage.value,
            "label": label,
            "wip_quantity": round(wip, 2)
        })
    
    return {
        "date": target_date.isoformat(),
        "wip_analysis": wip_data
    }

@app.get("/api/stages/config", response_model=List[dict])
def get_stage_configurations(db: Session = Depends(get_db)):
    """
    Get configuration for all production stages
    """
    configs = db.query(StageConfiguration).order_by(StageConfiguration.sequence_order).all()
    
    return [
        {
            "stage": config.stage.value,
            "sequence_order": config.sequence_order,
            "expected_input_size_mm": config.expected_input_size_mm,
            "expected_output_size_mm": config.expected_output_size_mm,
            "min_efficiency": config.min_efficiency,
            "max_loss_percentage": config.max_loss_percentage,
            "has_annealing": bool(config.has_annealing)
        }
        for config in configs
    ]

@app.get("/api/stats/summary")
def get_statistics_summary(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: Session = Depends(get_db)
):
    """
    Get comprehensive statistics summary for all stages
    """
    analytics = ProductionAnalytics(db)
    
    if not end_date:
        end_date = date.today()
    if not start_date:
        start_date = end_date - timedelta(days=30)
    
    stages_stats = []
    
    for stage in StageEnum:
        stats = analytics.get_stage_statistics(stage, start_date, end_date)
        stages_stats.append({
            "stage": stage.value,
            "total_input": stats.total_input,
            "total_output": stats.total_output,
            "total_scrap": stats.total_scrap,
            "avg_efficiency": stats.avg_efficiency,
            "avg_loss_percentage": stats.avg_loss_percentage,
            "record_count": stats.record_count
        })
    
    return {
        "date_range": {
            "start": start_date.isoformat(),
            "end": end_date.isoformat()
        },
        "stages": stages_stats
    }


# ==================== INVENTORY MANAGEMENT ENDPOINTS ====================

@app.get("/api/inventory/summary")
def get_inventory_summary(db: Session = Depends(get_db)):
    """
    Get overall inventory summary including total stock, stage-wise breakdown, and status
    """
    inventory_mgr = InventoryManager(db)
    return inventory_mgr.get_inventory_summary()


@app.get("/api/inventory/all")
def get_all_inventory(db: Session = Depends(get_db)):
    """
    Get current inventory levels for all stages with detailed information
    """
    inventory_mgr = InventoryManager(db)
    inventory_data = inventory_mgr.get_all_inventory()
    
    return [
        {
            "id": inv.id,
            "stage": inv.stage if isinstance(inv.stage, str) else inv.stage.value,
            "current_stock": inv.current_stock,
            "wire_size_mm": inv.wire_size_mm,
            "wire_size_swg": inv.wire_size_swg,
            "min_stock_level": inv.min_stock_level,
            "max_stock_level": inv.max_stock_level,
            "stock_status": inventory_mgr._get_stock_status(inv),
            "last_updated": inv.last_updated.isoformat() if inv.last_updated else None
        }
        for inv in inventory_data
    ]


@app.get("/api/inventory/stage/{stage}")
def get_stage_inventory(stage: StageEnum, db: Session = Depends(get_db)):
    """
    Get inventory details for a specific production stage
    """
    inventory_mgr = InventoryManager(db)
    inventory = inventory_mgr.get_stage_inventory(stage)
    
    if not inventory:
        raise HTTPException(status_code=404, detail=f"Inventory not found for stage: {stage.value}")
    
    stage_name = inventory.stage if isinstance(inventory.stage, str) else inventory.stage.value
    
    return {
        "id": inventory.id,
        "stage": stage_name,
        "current_stock": inventory.current_stock,
        "wire_size_mm": inventory.wire_size_mm,
        "wire_size_swg": inventory.wire_size_swg,
        "min_stock_level": inventory.min_stock_level,
        "max_stock_level": inventory.max_stock_level,
        "stock_status": inventory_mgr._get_stock_status(inventory),
        "last_updated": inventory.last_updated.isoformat() if inventory.last_updated else None
    }


@app.post("/api/inventory/update")
def update_inventory(
    stage: StageEnum,
    quantity: float,
    transaction_type: str,
    order_id: Optional[int] = None,
    notes: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    Update inventory for a stage with sequential production flow
    - Adds material to current stage
    - Automatically subtracts from previous stage (if not RBD)
    - Optionally links the entry to an order for progress tracking
    - RBD → Inter → Oven → DPC → Rewind
    """
    if transaction_type not in ['IN', 'OUT']:
        raise HTTPException(status_code=400, detail="transaction_type must be 'IN' or 'OUT'")
    
    # Define production sequence
    stage_sequence = {
        'Inter': 'RBD',
        'Oven': 'Inter',
        'DPC': 'Oven',
        'Rewind': 'DPC'
    }
    
    inventory_mgr = InventoryManager(db)
    
    try:
        # If adding material (IN) and not the first stage, subtract from previous
        if transaction_type == 'IN' and stage.value in stage_sequence:
            previous_stage = stage_sequence[stage.value]
            
            # Remove from previous stage
            try:
                inventory_mgr.update_inventory(
                    stage=previous_stage,
                    quantity=quantity,
                    transaction_type='OUT',
                    notes=f"Moved to {stage.value}"
                )
            except ValueError as e:
                raise HTTPException(
                    status_code=400, 
                    detail=str(e)
                )
        
        # Add to current stage
        updated_inventory = inventory_mgr.update_inventory(
            stage=stage,
            quantity=quantity,
            transaction_type=transaction_type,
            notes=notes
        )
        order_update = None

        if order_id and transaction_type == 'IN':
            order_update = _update_order_stage_progress(
                db=db,
                order_id=order_id,
                stage=stage,
                quantity=quantity
            )
        
        stage_name = updated_inventory.stage if isinstance(updated_inventory.stage, str) else updated_inventory.stage.value
        
        message = f"Inventory updated successfully for {stage.value}"
        if transaction_type == 'IN' and stage.value in stage_sequence:
            message += f" (moved from {stage_sequence[stage.value]})"
        
        return {
            "success": True,
            "stage": stage_name,
            "new_stock_level": updated_inventory.current_stock,
            "transaction_type": transaction_type,
            "quantity": quantity,
            "message": message,
            "order_update": order_update
        }
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/inventory/movement")
def record_material_movement(
    from_stage: StageEnum,
    to_stage: StageEnum,
    quantity: float,
    wire_size_mm: Optional[float] = None,
    wire_size_swg: Optional[str] = None,
    notes: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    Record material movement from one stage to another
    This reduces inventory in from_stage and increases inventory in to_stage
    """
    inventory_mgr = InventoryManager(db)
    
    try:
        movement = inventory_mgr.record_material_movement(
            from_stage=from_stage,
            to_stage=to_stage,
            quantity=quantity,
            wire_size_mm=wire_size_mm,
            wire_size_swg=wire_size_swg,
            notes=notes
        )
        
        return {
            "success": True,
            "movement_id": movement.id,
            "from_stage": from_stage.value,
            "to_stage": to_stage.value,
            "quantity": quantity,
            "movement_date": movement.movement_date.isoformat(),
            "message": f"Material movement recorded: {quantity} kg from {from_stage.value} to {to_stage.value}"
        }
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/api/inventory/alerts")
def get_inventory_alerts(db: Session = Depends(get_db)):
    """
    Get inventory alerts for stages with low or high stock levels
    """
    inventory_mgr = InventoryManager(db)
    alerts = inventory_mgr.get_inventory_alerts()
    
    return {
        "alert_count": len(alerts),
        "alerts": [
            {
                "stage": alert["stage"],
                "alert_type": alert["alert_type"],
                "current_stock": alert["current_stock"],
                "threshold": alert["threshold"],
                "message": alert["message"]
            }
            for alert in alerts
        ]
    }


@app.get("/api/inventory/transactions")
def get_transaction_history(
    stage: Optional[StageEnum] = None,
    limit: int = Query(50, ge=1, le=500),
    db: Session = Depends(get_db)
):
    """
    Get inventory transaction history, optionally filtered by stage
    """
    inventory_mgr = InventoryManager(db)
    transactions = inventory_mgr.get_transaction_history(stage=stage, limit=limit)
    
    return [
        {
            "id": trans.id,
            "stage": trans.stage if isinstance(trans.stage, str) else trans.stage.value,
            "transaction_type": trans.transaction_type,
            "quantity": trans.quantity,
            "stock_before": trans.stock_before,
            "stock_after": trans.stock_after,
            "notes": trans.notes,
            "timestamp": trans.timestamp.isoformat()
        }
        for trans in transactions
    ]


@app.get("/api/inventory/movements")
def get_material_movements(
    from_stage: Optional[StageEnum] = None,
    to_stage: Optional[StageEnum] = None,
    limit: int = Query(50, ge=1, le=500),
    db: Session = Depends(get_db)
):
    """
    Get material movement history between stages, optionally filtered
    """
    inventory_mgr = InventoryManager(db)
    movements = inventory_mgr.get_material_movements(
        from_stage=from_stage,
        to_stage=to_stage,
        limit=limit
    )
    
    return [
        {
            "id": mov.id,
            "from_stage": mov.from_stage if isinstance(mov.from_stage, str) else (mov.from_stage.value if mov.from_stage else None),
            "to_stage": mov.to_stage if isinstance(mov.to_stage, str) else (mov.to_stage.value if mov.to_stage else None),
            "quantity": mov.quantity,
            "wire_size_mm": mov.wire_size_mm,
            "wire_size_swg": mov.wire_size_swg,
            "movement_date": mov.movement_date.isoformat(),
            "notes": mov.notes
        }
        for mov in movements
    ]


@app.post("/api/inventory/sync")
def sync_inventory_from_production(
    start_date: Optional[date] = Query(None),
    db: Session = Depends(get_db)
):
    """
    Synchronize inventory levels based on production records
    Useful for recalculating inventory from production data
    """
    inventory_mgr = InventoryManager(db)
    
    try:
        inventory_mgr.sync_inventory_from_production(start_date=start_date)
        summary = inventory_mgr.get_inventory_summary()
        
        return {
            "success": True,
            "message": "Inventory synchronized with production records",
            "start_date": start_date.isoformat() if start_date else "beginning",
            "total_stock": summary["total_stock"],
            "stages_synced": len(summary["stages"])
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error syncing inventory: {str(e)}")


@app.get("/api/reports/production-summary")
def get_production_summary_report(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: Session = Depends(get_db)
):
    """Generate production and order summary for reporting/export"""
    default_end = end_date or date.today()
    default_start = start_date or (default_end - timedelta(days=30))

    records_query = db.query(ProductionRecord).filter(ProductionRecord.date >= default_start, ProductionRecord.date <= default_end)
    records = records_query.all()

    totals = {
        "total_input": 0.0,
        "total_output": 0.0,
        "total_scrap": 0.0
    }

    stage_summary = defaultdict(lambda: {"input": 0.0, "output": 0.0, "scrap": 0.0})

    for record in records:
        stage_name = record.stage.value if isinstance(record.stage, StageEnum) else record.stage
        totals["total_input"] += record.input_qty
        totals["total_output"] += record.output_qty
        totals["total_scrap"] += record.scrap_qty
        stage_summary[stage_name]["input"] += record.input_qty
        stage_summary[stage_name]["output"] += record.output_qty
        stage_summary[stage_name]["scrap"] += record.scrap_qty

    stage_breakdown = []
    for stage_name, data in stage_summary.items():
        efficiency = (data["output"] / data["input"] * 100) if data["input"] > 0 else 0
        stage_breakdown.append({
            "stage": stage_name,
            "input": round(data["input"], 2),
            "output": round(data["output"], 2),
            "scrap": round(data["scrap"], 2),
            "efficiency": round(efficiency, 2)
        })

    orders = db.query(ProductionOrder).all()
    order_summary = {
        "total_orders": len(orders),
        "pending": 0,
        "in_progress": 0,
        "completed": 0,
        "by_stage": defaultdict(int)
    }

    for order in orders:
        if order.status == OrderStatus.COMPLETED.value:
            order_summary["completed"] += 1
        elif order.status == OrderStatus.IN_PROGRESS.value:
            order_summary["in_progress"] += 1
        else:
            order_summary["pending"] += 1

        if order.current_stage:
            order_summary["by_stage"][order.current_stage] += 1

    utilization = (totals["total_output"] / totals["total_input"] * 100) if totals["total_input"] > 0 else 0

    return {
        "date_range": {
            "start": default_start.isoformat(),
            "end": default_end.isoformat()
        },
        "totals": {
            **{k: round(v, 2) for k, v in totals.items()},
            "utilization": round(utilization, 2)
        },
        "stage_breakdown": sorted(stage_breakdown, key=lambda item: item["stage"]),
        "order_summary": {
            "total_orders": order_summary["total_orders"],
            "pending": order_summary["pending"],
            "in_progress": order_summary["in_progress"],
            "completed": order_summary["completed"],
            "by_stage": dict(order_summary["by_stage"])
        }
    }


@app.get("/api/reports/inventory-transactions/export")
def export_inventory_transactions(
    stage: Optional[StageEnum] = None,
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: Session = Depends(get_db)
):
    """Export inventory transactions as Excel file"""
    query = db.query(InventoryTransaction)

    if stage:
        query = query.filter(InventoryTransaction.stage == stage.value)

    if start_date:
        query = query.filter(InventoryTransaction.timestamp >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        query = query.filter(InventoryTransaction.timestamp <= datetime.combine(end_date, datetime.max.time()))

    transactions = query.order_by(InventoryTransaction.timestamp.desc()).all()

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Transactions"
    
    # Define header style
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    headers = ["Transaction ID", "Stage", "Type", "Quantity (kg)", "Stock Before", "Stock After", "Notes", "Timestamp"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write data
    for row_num, trans in enumerate(transactions, 2):
        ws.cell(row=row_num, column=1, value=trans.id)
        ws.cell(row=row_num, column=2, value=trans.stage)
        ws.cell(row=row_num, column=3, value=trans.transaction_type)
        ws.cell(row=row_num, column=4, value=round(trans.quantity, 2))
        ws.cell(row=row_num, column=5, value=round(trans.stock_before, 2))
        ws.cell(row=row_num, column=6, value=round(trans.stock_after, 2))
        ws.cell(row=row_num, column=7, value=trans.notes or "")
        ws.cell(row=row_num, column=8, value=trans.timestamp.isoformat() if trans.timestamp else "")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"inventory_transactions_{date.today().isoformat()}.xlsx"
    headers = {
        "Content-Disposition": f"attachment; filename={filename}"
    }
    return StreamingResponse(
        iter([output.getvalue()]), 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers=headers
    )


# ==================== ORDER TRACKING ENDPOINTS ====================




@app.get("/api/orders")
def get_all_orders(
    status: Optional[str] = None,
    customer: Optional[str] = None,
    limit: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db)
):
    """
    Get all production orders with optional filtering
    """
    query = db.query(ProductionOrder)
    
    if status:
        query = query.filter(ProductionOrder.status == status)
    
    if customer:
        query = query.filter(ProductionOrder.customer_name.ilike(f"%{customer}%"))
    
    orders = query.order_by(ProductionOrder.created_at.desc()).limit(limit).all()
    
    return [
        {
            "id": order.id,
            "order_number": order.order_number,
            "customer_name": order.customer_name,
            "product_specification": order.product_specification,
            "ordered_quantity": order.ordered_quantity,
            "completed_quantity": order.completed_quantity,
            "status": order.status,
            "current_stage": order.current_stage,
            "order_date": order.order_date.isoformat(),
            "expected_delivery_date": order.expected_delivery_date.isoformat() if order.expected_delivery_date else None,
            "actual_delivery_date": order.actual_delivery_date.isoformat() if order.actual_delivery_date else None,
            "priority": order.priority,
            "completion_percentage": round((order.completed_quantity / order.ordered_quantity * 100), 2) if order.ordered_quantity > 0 else 0
        }
        for order in orders
    ]


@app.post("/api/orders")
def create_order(payload: OrderCreatePayload = Body(...), db: Session = Depends(get_db)):
    """Create a new production order"""
    try:
        existing = db.query(ProductionOrder).filter(ProductionOrder.order_number == payload.order_number).first()
        if existing:
            raise HTTPException(status_code=400, detail=f"Order {payload.order_number} already exists")

        order = ProductionOrder(
            order_number=payload.order_number,
            customer_name=payload.customer_name,
            product_specification=payload.product_specification,
            ordered_quantity=payload.ordered_quantity,
            target_wire_size_mm=payload.target_wire_size_mm,
            expected_delivery_date=payload.expected_delivery_date,
            priority=payload.priority,
            notes=payload.notes,
            order_date=date.today(),
            status=OrderStatus.PENDING.value,
            completed_quantity=0.0
        )
        
        db.add(order)
        db.commit()
        db.refresh(order)
        
        # Initialize stage progress for all stages
        for stage in StageEnum:
            stage_progress = OrderStageProgress(
                order_id=order.id,
                stage=stage.value,
                stage_status="PENDING"
            )
            db.add(stage_progress)
        
        db.commit()
        
        return {
            "success": True,
            "message": f"Order {order.order_number} created successfully",
            "order": {
                "id": order.id,
                "order_number": order.order_number,
                "customer_name": order.customer_name,
                "status": order.status
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"ERROR creating order: {str(e)}")
        import traceback
        traceback.print_exc()
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to create order: {str(e)}")

@app.get("/api/orders/{order_id}")
def get_order_details(order_id: int, db: Session = Depends(get_db)):
    """
    Get detailed information about a specific order including stage progress
    """
    order = db.query(ProductionOrder).filter(ProductionOrder.id == order_id).first()
    
    if not order:
        raise HTTPException(status_code=404, detail=f"Order {order_id} not found")
    
    # Get stage progress
    stage_progress = db.query(OrderStageProgress).filter(
        OrderStageProgress.order_id == order_id
    ).all()
    
    return {
        "id": order.id,
        "order_number": order.order_number,
        "customer_name": order.customer_name,
        "product_specification": order.product_specification,
        "target_wire_size_mm": order.target_wire_size_mm,
        "target_wire_size_swg": order.target_wire_size_swg,
        "ordered_quantity": order.ordered_quantity,
        "completed_quantity": order.completed_quantity,
        "status": order.status,
        "current_stage": order.current_stage,
        "order_date": order.order_date.isoformat(),
        "expected_delivery_date": order.expected_delivery_date.isoformat() if order.expected_delivery_date else None,
        "actual_delivery_date": order.actual_delivery_date.isoformat() if order.actual_delivery_date else None,
        "priority": order.priority,
        "notes": order.notes,
        "completion_percentage": round((order.completed_quantity / order.ordered_quantity * 100), 2) if order.ordered_quantity > 0 else 0,
        "stage_progress": [
            {
                "stage": sp.stage,
                "input_quantity": sp.input_quantity,
                "output_quantity": sp.output_quantity,
                "scrap_quantity": sp.scrap_quantity,
                "stage_status": sp.stage_status,
                "started_at": sp.started_at.isoformat() if sp.started_at else None,
                "completed_at": sp.completed_at.isoformat() if sp.completed_at else None
            }
            for sp in stage_progress
        ]
    }


@app.put("/api/orders/{order_id}/status")
def update_order_status(
    order_id: int,
    status: str,
    current_stage: Optional[str] = None,
    completed_quantity: Optional[float] = None,
    actual_delivery_date: Optional[date] = None,
    db: Session = Depends(get_db)
):
    """
    Update order status and progress
    """
    order = db.query(ProductionOrder).filter(ProductionOrder.id == order_id).first()
    
    if not order:
        raise HTTPException(status_code=404, detail=f"Order {order_id} not found")
    
    order.status = status
    if current_stage:
        order.current_stage = current_stage
    if completed_quantity is not None:
        order.completed_quantity = completed_quantity
    if actual_delivery_date:
        order.actual_delivery_date = actual_delivery_date
    
    db.commit()
    db.refresh(order)
    
    return {
        "success": True,
        "order_id": order.id,
        "order_number": order.order_number,
        "status": order.status,
        "message": "Order updated successfully"
    }


# ==================== BATCH TRACKING ENDPOINTS ====================

@app.get("/api/batches")
def list_batches(
    status: Optional[str] = None,
    stage: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Return all tracked coils/batches with latest journey context"""
    stage_sequence = _stage_sequence()
    query = db.query(BatchTracking)

    if status:
        query = query.filter(BatchTracking.current_status == status)
    if stage:
        query = query.filter(BatchTracking.current_stage == stage)

    batches = query.order_by(BatchTracking.updated_at.desc()).all()
    order_lookup = {}
    order_ids = {batch.order_id for batch in batches if batch.order_id}
    if order_ids:
        orders = db.query(ProductionOrder).filter(ProductionOrder.id.in_(order_ids)).all()
        order_lookup = {order.id: order for order in orders}

    response = []
    for batch in batches:
        last_event = db.query(BatchJourneyEvent).filter(
            BatchJourneyEvent.batch_id == batch.id
        ).order_by(BatchJourneyEvent.movement_date.desc()).first()
        response.append(_serialize_batch(batch, stage_sequence, order_lookup, last_event))

    return response


@app.get("/api/batches/{batch_id}")
def get_batch_detail(batch_id: int, db: Session = Depends(get_db)):
    """Detailed journey + metadata for a specific batch"""
    batch = db.query(BatchTracking).filter(BatchTracking.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail=f"Batch {batch_id} not found")

    stage_sequence = _stage_sequence()
    order = db.query(ProductionOrder).filter(ProductionOrder.id == batch.order_id).first() if batch.order_id else None
    events = db.query(BatchJourneyEvent).filter(BatchJourneyEvent.batch_id == batch_id).order_by(BatchJourneyEvent.movement_date.asc()).all()
    order_lookup = {order.id: order} if order else {}

    detail = _serialize_batch(batch, stage_sequence, order_lookup)
    detail.update({
        "journey_events": [_serialize_journey_event(evt) for evt in events],
        "next_stage": _get_next_stage(batch.current_stage, stage_sequence)
    })
    if order:
        detail["order"] = {
            "order_number": order.order_number,
            "customer_name": order.customer_name,
            "status": order.status,
            "ordered_quantity": order.ordered_quantity,
            "completed_quantity": order.completed_quantity
        }

    return detail


@app.post("/api/batches")
def create_batch(payload: BatchCreatePayload, db: Session = Depends(get_db)):
    """Register a new coil/batch for traceability"""
    try:
        existing = db.query(BatchTracking).filter(BatchTracking.batch_number == payload.batch_number).first()
        if existing:
            raise HTTPException(status_code=400, detail=f"Batch {payload.batch_number} already exists")

        initial_stage = payload.initial_stage.value if payload.initial_stage else None
        batch = BatchTracking(
            batch_number=payload.batch_number,
            lot_number=payload.lot_number,
            material_type=payload.material_type,
            quantity=payload.quantity,
            remaining_quantity=payload.quantity,
            current_stage=initial_stage,
            current_status="ACTIVE",
            supplier_name=payload.supplier_name,
            received_date=payload.received_date,
            order_id=payload.order_id,
            notes=payload.notes
        )
        db.add(batch)
        db.commit()
        db.refresh(batch)

        # Sync inventory
        if initial_stage:
            try:
                inventory_mgr = InventoryManager(db)
                inventory_mgr.update_inventory(
                    stage=initial_stage,
                    quantity=payload.quantity,
                    transaction_type='IN',
                    notes=f"Batch {batch.batch_number} created"
                )
            except Exception as e:
                # Log error but don't fail batch creation? 
                # Or fail? Better to fail or at least log.
                print(f"Failed to update inventory for batch {batch.batch_number}: {e}")

        stage_sequence = _stage_sequence()
        order_lookup = {}
        if batch.order_id:
            order = db.query(ProductionOrder).filter(ProductionOrder.id == batch.order_id).first()
            if order:
                order_lookup[order.id] = order

        return {
            "success": True,
            "message": f"Batch {batch.batch_number} registered",
            "batch": _serialize_batch(batch, stage_sequence, order_lookup)
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"ERROR creating batch: {str(e)}")
        import traceback
        traceback.print_exc()
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to create batch: {str(e)}")


@app.post("/api/batches/{batch_id}/move")
def move_batch_forward(
    batch_id: int,
    payload: BatchMovePayload,
    db: Session = Depends(get_db)
):
    """Log a guided move for a batch and advance it to the next stage"""
    batch = db.query(BatchTracking).filter(BatchTracking.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail=f"Batch {batch_id} not found")

    stage_sequence = _stage_sequence()
    target_stage = payload.to_stage.value
    previous_stage = batch.current_stage

    if target_stage not in stage_sequence:
        raise HTTPException(status_code=400, detail=f"Stage {target_stage} is not part of the configured flow")

    if previous_stage:
        if previous_stage not in stage_sequence:
            raise HTTPException(status_code=400, detail=f"Existing stage {previous_stage} is invalid")
        prev_idx = stage_sequence.index(previous_stage)
        target_idx = stage_sequence.index(target_stage)
        if target_idx < prev_idx:
            raise HTTPException(status_code=400, detail="Cannot move backwards in the process")
        if target_idx - prev_idx > 1:
            raise HTTPException(status_code=400, detail="Moves must follow the guided stage order")
    else:
        first_stage = stage_sequence[0] if stage_sequence else None
        if first_stage and target_stage != first_stage:
            raise HTTPException(status_code=400, detail=f"First move must enter {first_stage}")

    if batch.remaining_quantity is None:
        batch.remaining_quantity = batch.quantity

    available_qty = batch.remaining_quantity
    if available_qty <= 0:
        raise HTTPException(status_code=400, detail=f"Batch {batch.batch_number} has no remaining quantity to move")

    if payload.scrap_quantity > payload.quantity:
        raise HTTPException(status_code=400, detail="Scrap quantity cannot exceed moved quantity")

    if payload.quantity > available_qty:
        raise HTTPException(status_code=400, detail=f"Cannot move {payload.quantity} kg; only {available_qty} kg remaining on batch {batch.batch_number}")

    if payload.scrap_quantity > available_qty:
        raise HTTPException(status_code=400, detail=f"Cannot scrap {payload.scrap_quantity} kg; only {available_qty} kg remaining on batch {batch.batch_number}")

    event = BatchJourneyEvent(
        batch_id=batch.id,
        from_stage=previous_stage,
        to_stage=target_stage,
        quantity=payload.quantity,
        scrap_quantity=payload.scrap_quantity,
        operator=payload.operator,
        notes=payload.notes
    )
    db.add(event)

    inventory_mgr = InventoryManager(db)
    try:
        inventory_mgr.record_material_movement(
            from_stage=previous_stage,
            to_stage=target_stage,
            quantity=payload.quantity,
            notes=f"Batch {batch.batch_number}: {payload.notes or 'Guided move'}",
            batch_id=batch.id,
            batch_number=batch.batch_number
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    batch.current_stage = target_stage
    batch.current_status = "CONSUMED" if target_stage == StageEnum.REWIND.value else "ACTIVE"
    batch.remaining_quantity = max(0.0, available_qty - payload.scrap_quantity)

    db.commit()
    db.refresh(batch)
    db.refresh(event)

    order_lookup = {}
    order_update = None
    if batch.order_id:
        order = db.query(ProductionOrder).filter(ProductionOrder.id == batch.order_id).first()
        if order:
            order_lookup[order.id] = order
            if payload.to_stage == StageEnum.REWIND:
                order_update = _update_order_stage_progress(db, batch.order_id, payload.to_stage, payload.quantity)
                
    # Create production record for Rewind stage to update summary cards (regardless of order linkage)
    if payload.to_stage == StageEnum.REWIND:
        production_record = ProductionRecord(
            date=date.today(),
            shift=ShiftEnum.MORNING,  # Default to morning shift
            stage=StageEnum.REWIND,
            input_qty=payload.quantity + payload.scrap_quantity,
            output_qty=payload.quantity,
            scrap_qty=payload.scrap_quantity,
            efficiency=(payload.quantity / (payload.quantity + payload.scrap_quantity) * 100) if (payload.quantity + payload.scrap_quantity) > 0 else 0,
            loss_percentage=(payload.scrap_quantity / (payload.quantity + payload.scrap_quantity) * 100) if (payload.quantity + payload.scrap_quantity) > 0 else 0,
            notes=f"Batch {batch.batch_number} completed"
        )
        db.add(production_record)
        db.commit()

    stage_sequence = _stage_sequence()

    return {
        "success": True,
        "message": f"Batch {batch.batch_number} moved to {target_stage}",
        "batch": _serialize_batch(batch, stage_sequence, order_lookup, event),
        "order_update": order_update,
        "next_stage": _get_next_stage(batch.current_stage, stage_sequence)
    }


@app.post("/api/batches/{batch_id}/hold")
def toggle_batch_hold(
    batch_id: int,
    payload: BatchHoldPayload,
    db: Session = Depends(get_db)
):
    """Pause or resume a batch mid-journey"""
    batch = db.query(BatchTracking).filter(BatchTracking.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail=f"Batch {batch_id} not found")

    if batch.current_status == "CONSUMED" and payload.hold:
        raise HTTPException(status_code=400, detail="Completed batches cannot be placed on hold")

    target_status = "ON_HOLD" if payload.hold else "ACTIVE"
    batch.current_status = target_status if batch.current_status != "CONSUMED" else batch.current_status

    stage_marker = batch.current_stage or "UNASSIGNED"
    reason_text = payload.reason.strip() if payload.reason else None
    action_phrase = "HOLD" if payload.hold else "RESUME"
    event_note = reason_text or ("Batch paused" if payload.hold else "Batch resumed")

    journey_event = BatchJourneyEvent(
        batch_id=batch.id,
        from_stage=stage_marker if payload.hold else "HOLD",
        to_stage="HOLD" if payload.hold else stage_marker,
        quantity=0.0,
        scrap_quantity=0.0,
        notes=f"{action_phrase.title()}: {event_note}"
    )
    db.add(journey_event)

    hold_history = _parse_hold_history(batch.hold_history or batch.quality_notes)
    hold_history.append({
        "action": action_phrase,
        "reason": reason_text,
        "timestamp": datetime.utcnow().isoformat()
    })
    batch.hold_history = json.dumps(hold_history)

    db.commit()
    db.refresh(batch)

    return {
        "success": True,
        "batch_id": batch.id,
        "batch_number": batch.batch_number,
        "status": batch.current_status,
        "message": "Batch placed on hold" if payload.hold else "Batch resumed"
    }


# ==================== PRODUCTION ENTRY ENDPOINTS ====================

@app.post("/api/production/entry")
def create_production_entry(
    date: date,
    shift: ShiftEnum,
    stage: StageEnum,
    input_qty: float,
    output_qty: float,
    scrap_qty: float = 0.0,
    input_size_mm: Optional[float] = None,
    output_size_mm: Optional[float] = None,
    input_size_swg: Optional[int] = None,
    output_size_swg: Optional[int] = None,
    remarks: Optional[str] = None,
    order_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """
    Create a new production record - replaces manual Excel entry
    """
    # Calculate efficiency and loss
    efficiency = (output_qty / input_qty * 100) if input_qty > 0 else 0
    loss_percentage = (scrap_qty / input_qty * 100) if input_qty > 0 else 0
    
    # Create production record
    record = ProductionRecord(
        date=date,
        shift=shift,
        stage=stage,
        input_qty=input_qty,
        output_qty=output_qty,
        scrap_qty=scrap_qty,
        input_size_mm=input_size_mm,
        output_size_mm=output_size_mm,
        input_size_swg=input_size_swg,
        output_size_swg=output_size_swg,
        efficiency=efficiency,
        loss_percentage=loss_percentage,
        remarks=remarks
    )
    
    db.add(record)
    db.commit()
    db.refresh(record)
    
    # Update inventory
    inventory_mgr = InventoryManager(db)
    inventory_mgr.update_inventory(
        stage=stage,
        quantity=output_qty,
        transaction_type='IN',
        notes=f"Production entry: {date} {shift.value}"
    )
    
    # If linked to an order, update order progress
    if order_id:
        order = db.query(ProductionOrder).filter(ProductionOrder.id == order_id).first()
        if order:
            # Update stage progress
            stage_progress = db.query(OrderStageProgress).filter(
                OrderStageProgress.order_id == order_id,
                OrderStageProgress.stage == stage.value
            ).first()
            
            if stage_progress:
                stage_progress.input_quantity += input_qty
                stage_progress.output_quantity += output_qty
                stage_progress.scrap_quantity += scrap_qty
                stage_progress.production_record_id = record.id
                stage_progress.stage_status = "IN_PROGRESS"
                
                if not stage_progress.started_at:
                    from datetime import datetime
                    stage_progress.started_at = datetime.now()
                
                db.commit()
            
            # Update order current stage and status
            order.current_stage = stage.value
            if order.status == OrderStatus.PENDING.value:
                order.status = OrderStatus.IN_PROGRESS.value
            
            # If this is final stage (REWIND), update completed quantity
            if stage == StageEnum.REWIND:
                order.completed_quantity += output_qty
            
            db.commit()
    
    return {
        "success": True,
        "record_id": record.id,
        "efficiency": round(efficiency, 2),
        "loss_percentage": round(loss_percentage, 2),
        "message": "Production entry recorded successfully"
    }


@app.get("/api/production/quick-stats")
def get_quick_stats(db: Session = Depends(get_db)):
    """
    Get quick statistics for production entry screen
    """
    today = date.today()
    
    # Today's production records
    today_records = db.query(ProductionRecord).filter(
        ProductionRecord.date == today
    ).all()
    
    # Active orders
    active_orders = db.query(ProductionOrder).filter(
        ProductionOrder.status.in_([OrderStatus.PENDING.value, OrderStatus.IN_PROGRESS.value])
    ).count()
    
    # Current inventory summary
    inventory_mgr = InventoryManager(db)
    inventory_summary = inventory_mgr.get_inventory_summary()
    
    return {
        "today_entries": len(today_records),
        "active_orders": active_orders,
        "total_inventory": inventory_summary["total_stock"],
        "low_stock_alerts": inventory_summary["low_stock_count"],
        "last_entry_time": today_records[-1].created_at.isoformat() if today_records else None
    }


# ==================== REPORT GENERATION ENDPOINTS ====================

@app.get("/api/reports/production-summary")
def get_production_summary_report(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: Session = Depends(get_db)
):
    """
    Generate production summary report for a date range
    """
    if not end_date:
        end_date = date.today()
    if not start_date:
        start_date = end_date - timedelta(days=30)
    
    analytics = ProductionAnalytics(db)
    
    # Get overall metrics
    overall = analytics.get_overall_metrics(start_date, end_date)
    
    # Get stage-wise statistics
    stage_stats = []
    for stage in StageEnum:
        stats = analytics.get_stage_statistics(stage, start_date, end_date)
        stage_stats.append({
            "stage": stage.value,
            "total_input": stats.total_input,
            "total_output": stats.total_output,
            "total_scrap": stats.total_scrap,
            "avg_efficiency": stats.avg_efficiency,
            "avg_loss_percentage": stats.avg_loss_percentage,
            "record_count": stats.record_count
        })
    
    # Get inventory summary
    inventory_mgr = InventoryManager(db)
    inventory_summary = inventory_mgr.get_inventory_summary()
    
    # Get alerts
    alerts = analytics.generate_alerts(start_date, end_date)
    
    return {
        "report_generated_at": date.today().isoformat(),
        "date_range": {
            "start": start_date.isoformat(),
            "end": end_date.isoformat(),
            "days": (end_date - start_date).days
        },
        "overall_metrics": {
            "total_production": overall.total_production,
            "total_efficiency": overall.total_efficiency,
            "total_scrap": overall.total_scrap,
            "scrap_percentage": overall.scrap_percentage
        },
        "stage_statistics": stage_stats,
        "inventory_summary": inventory_summary,
        "alerts_count": len(alerts),
        "critical_alerts": len([a for a in alerts if a["severity"] == "CRITICAL"])
    }


@app.get("/api/reports/order-status")
def get_order_status_report(
    status: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    Generate report of all orders with their status
    """
    query = db.query(ProductionOrder)
    
    if status:
        query = query.filter(ProductionOrder.status == status)
    
    orders = query.all()
    
    # Calculate statistics
    total_ordered = sum(o.ordered_quantity for o in orders)
    total_completed = sum(o.completed_quantity for o in orders)
    
    status_counts = {
        "PENDING": len([o for o in orders if o.status == OrderStatus.PENDING.value]),
        "IN_PROGRESS": len([o for o in orders if o.status == OrderStatus.IN_PROGRESS.value]),
        "COMPLETED": len([o for o in orders if o.status == OrderStatus.COMPLETED.value]),
        "ON_HOLD": len([o for o in orders if o.status == OrderStatus.ON_HOLD.value]),
        "CANCELLED": len([o for o in orders if o.status == OrderStatus.CANCELLED.value])
    }
    
    # Orders by priority
    urgent_orders = len([o for o in orders if o.priority == 3 and o.status != OrderStatus.COMPLETED.value])
    high_priority = len([o for o in orders if o.priority == 2 and o.status != OrderStatus.COMPLETED.value])
    
    return {
        "report_generated_at": date.today().isoformat(),
        "total_orders": len(orders),
        "total_ordered_quantity": round(total_ordered, 2),
        "total_completed_quantity": round(total_completed, 2),
        "overall_completion_percentage": round((total_completed / total_ordered * 100), 2) if total_ordered > 0 else 0,
        "status_breakdown": status_counts,
        "urgent_orders": urgent_orders,
        "high_priority_orders": high_priority,
        "orders": [
            {
                "order_number": o.order_number,
                "customer_name": o.customer_name,
                "status": o.status,
                "ordered_quantity": o.ordered_quantity,
                "completed_quantity": o.completed_quantity,
                "completion_percentage": round((o.completed_quantity / o.ordered_quantity * 100), 2) if o.ordered_quantity > 0 else 0,
                "order_date": o.order_date.isoformat(),
                "expected_delivery_date": o.expected_delivery_date.isoformat() if o.expected_delivery_date else None,
                "priority": o.priority
            }
            for o in orders
        ]
    }

@app.get("/api/reports/production/export")
def export_production_report(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: Session = Depends(get_db)
):
    """
    Export production records as formatted Excel
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import Response

    if not end_date:
        end_date = date.today()
    if not start_date:
        start_date = end_date - timedelta(days=30)
    
    records = db.query(ProductionRecord).filter(
        ProductionRecord.date >= start_date,
        ProductionRecord.date <= end_date
    ).order_by(ProductionRecord.date.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Production Report"
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    headers = [
        "Date", "Shift", "Stage", "Input (kg)", "Output (kg)", 
        "Scrap (kg)", "Efficiency (%)", "Loss (%)", "Notes"
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        
    # Write data
    for row_num, record in enumerate(records, 2):
        ws.cell(row=row_num, column=1, value=record.date)
        ws.cell(row=row_num, column=2, value=record.shift.value)
        ws.cell(row=row_num, column=3, value=record.stage.value)
        ws.cell(row=row_num, column=4, value=record.input_qty)
        ws.cell(row=row_num, column=5, value=record.output_qty)
        ws.cell(row=row_num, column=6, value=record.scrap_qty)
        ws.cell(row=row_num, column=7, value=record.efficiency)
        ws.cell(row=row_num, column=8, value=record.loss_percentage)
        ws.cell(row=row_num, column=9, value=record.notes or "")
        
    # Auto-adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value) or "") for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=production_report_{start_date}_{end_date}.xlsx"
        }
    )

@app.get("/api/reports/production/export/pdf")
def export_production_report_pdf(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: Session = Depends(get_db)
):
    """
    Export production records as PDF with Logo
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from fastapi.responses import Response
    import os

    if not end_date:
        end_date = date.today()
    if not start_date:
        start_date = end_date - timedelta(days=30)
    
    records = db.query(ProductionRecord).filter(
        ProductionRecord.date >= start_date,
        ProductionRecord.date <= end_date
    ).order_by(ProductionRecord.date.desc()).all()
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output, 
        pagesize=landscape(letter),
        rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Add Logo
    logo_path = os.path.join(os.getcwd(), "../frontend/public/tulsi-logo.png")
    if os.path.exists(logo_path):
        im = Image(logo_path, width=120, height=40)
        im.hAlign = 'LEFT'
        elements.append(im)
        elements.append(Spacer(1, 20))
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30
    )
    elements.append(Paragraph(f"Production Report ({start_date} to {end_date})", title_style))
    
    # Table Data
    data = [["Date", "Shift", "Stage", "Input", "Output", "Scrap", "Eff %", "Loss %", "Notes"]]
    
    for record in records:
        data.append([
            str(record.date),
            record.shift.value,
            record.stage.value,
            f"{record.input_qty:.1f}",
            f"{record.output_qty:.1f}",
            f"{record.scrap_qty:.1f}",
            f"{record.efficiency:.1f}%",
            f"{record.loss_percentage:.1f}%",
            record.notes[:30] + "..." if record.notes and len(record.notes) > 30 else (record.notes or "")
        ])
        
    # Table Style
    table = Table(data, colWidths=[70, 60, 70, 60, 60, 60, 60, 60, 150])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
        ('ALIGN', (-1, 0), (-1, -1), 'LEFT'), # Align notes left
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    return Response(
        content=output.getvalue(),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=production_report_{start_date}_{end_date}.pdf"
        }
    )

@app.get("/api/reports/efficiency-analysis")
def get_efficiency_analysis_report(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: Session = Depends(get_db)
):
    """
    Generate efficiency analysis report by stage and shift
    """
    if not end_date:
        end_date = date.today()
    if not start_date:
        start_date = end_date - timedelta(days=30)
    
    records = db.query(ProductionRecord).filter(
        ProductionRecord.date >= start_date,
        ProductionRecord.date <= end_date
    ).all()
    
    # Analysis by stage
    stage_analysis = {}
    for stage in StageEnum:
        stage_records = [r for r in records if r.stage == stage]
        if stage_records:
            stage_analysis[stage.value] = {
                "avg_efficiency": round(sum(r.efficiency for r in stage_records) / len(stage_records), 2),
                "min_efficiency": round(min(r.efficiency for r in stage_records), 2),
                "max_efficiency": round(max(r.efficiency for r in stage_records), 2),
                "avg_loss": round(sum(r.loss_percentage for r in stage_records) / len(stage_records), 2),
                "total_records": len(stage_records)
            }
    
    # Analysis by shift
    shift_analysis = {}
    for shift in ShiftEnum:
        shift_records = [r for r in records if r.shift == shift]
        if shift_records:
            shift_analysis[shift.value] = {
                "avg_efficiency": round(sum(r.efficiency for r in shift_records) / len(shift_records), 2),
                "avg_loss": round(sum(r.loss_percentage for r in shift_records) / len(shift_records), 2),
                "total_output": round(sum(r.output_qty for r in shift_records), 2),
                "total_scrap": round(sum(r.scrap_qty for r in shift_records), 2)
            }
    
    # Find bottleneck stage
    analytics = ProductionAnalytics(db)
    bottleneck = analytics.detect_bottleneck(start_date, end_date)
    
    return {
        "report_generated_at": date.today().isoformat(),
        "date_range": {
            "start": start_date.isoformat(),
            "end": end_date.isoformat()
        },
        "stage_analysis": stage_analysis,
        "shift_analysis": shift_analysis,
        "bottleneck_stage": bottleneck["stage"] if bottleneck else None,
        "total_records_analyzed": len(records)
    }


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)

# ==================== FINAL STAGES ROUTES ====================
# Include final stage routes (Quality Check, Packaging, Dispatch)
from final_stage_routes import router as final_stage_router
app.include_router(final_stage_router, prefix="/api", tags=["Final Stages"])
